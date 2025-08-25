import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import plotly.colors as pc
from plotly.colors import hex_to_rgb
import plotly.io as pio

import polars as pl
import pandas as pd
import numpy as np
from datetime import datetime
import io

from data_simulation import simulate_ecommerce_data
from data_processing import clean_and_merge_data, prepare_model_data, process_uploaded_files

from models import PriceElasticityModels

# Set page config FIRST (must be the first Streamlit command)
st.set_page_config(page_title="Price Elasticity Analytics", layout="wide")

# Check if theme config exists and show current theme
if st.sidebar.checkbox("Show Theme Info", value=False):
    st.sidebar.info(f"""
    **Current Theme Settings:**
    - Primary Color: {st.get_option('theme.primaryColor') or 'Not set'}
    - Background: {st.get_option('theme.backgroundColor') or 'Not set'}
    - Text Color: {st.get_option('theme.textColor') or 'Not set'}
    
    To use custom theme, ensure `.streamlit/config.toml` exists in your app directory.
    """)

# Color utility functions
def get_primary_color():
    """Get primary color with fallback"""
    # Try to get from session state first (user override)
    if 'primary_color_override' in st.session_state:
        return st.session_state.primary_color_override
    
    # Then try config file
    config_color = st.get_option('theme.primaryColor')
    if config_color:
        return config_color
    
    # Default fallback
    return '#FF4B4B'  # Streamlit's default red

def create_color_palette(primary_color, n_colors=10):
    """Create a color palette based on the primary color"""
    try:
        rgb = hex_to_rgb(primary_color)
    except:
        # Fallback if color parsing fails
        return px.colors.qualitative.Plotly[:n_colors]
    
    colors = []
    colors.append(primary_color)
    
    # Create lighter versions
    for i in range(1, n_colors//2):
        factor = 1 + (i * 0.2)
        new_rgb = tuple(min(255, int(c + (255-c) * i * 0.2)) for c in rgb)
        colors.append(f'rgb{new_rgb}')
    
    # Create darker versions
    for i in range(1, n_colors//2 + 1):
        factor = 1 - (i * 0.15)
        new_rgb = tuple(max(0, int(c * factor)) for c in rgb)
        colors.append(f'rgb{new_rgb}')
    
    return colors[:n_colors]

def create_diverging_palette(primary_color):
    """Create a diverging color palette with primary color in the middle"""
    try:
        rgb = hex_to_rgb(primary_color)
        return [
            'rgb(215, 48, 39)',    # Red for low values
            f'rgb{rgb}',           # Primary color for middle
            'rgb(69, 182, 69)'     # Green for high values
        ]
    except:
        return ['red', primary_color, 'green']

def create_monochrome_continuous_scale(primary_color):
    """Create a continuous color scale based on primary color"""
    try:
        rgb = hex_to_rgb(primary_color)
        return [
            [0, f'rgb(255, 255, 255)'],  # White
            [0.5, primary_color],         # Primary color
            [1, f'rgb({int(rgb[0]*0.3)}, {int(rgb[1]*0.3)}, {int(rgb[2]*0.3)})']  # Dark version
        ]
    except:
        return [[0, 'white'], [0.5, primary_color], [1, 'black']]

# Initialize color settings
def get_current_colors():
    """Get current theme colors dynamically"""
    primary_color = get_primary_color()
    return {
        'primary': primary_color,
        'palette': create_color_palette(primary_color),
        'diverging': create_diverging_palette(primary_color),
        'continuous': create_monochrome_continuous_scale(primary_color)
    }

def style_plotly_fig(fig, use_primary_color=True):
    """Apply consistent styling to a Plotly figure using current theme"""
    
    # Get current colors
    colors = get_current_colors()
    primary_color = colors['primary']
    
    # Update layout to be transparent and work with Streamlit themes
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',  # Transparent background
        plot_bgcolor='rgba(0,0,0,0)',   # Transparent plot background
        font=dict(color=None),          # Let Streamlit handle font color
        title_x=0.5,
        margin=dict(l=20, r=20, t=40, b=20),
        hoverlabel=dict(
            bgcolor=primary_color,
            font_color="white",
            font_size=14
        ),
        colorway=colors['palette']
    )
    
    # Update axes with subtle grid
    fig.update_xaxes(
        gridcolor='rgba(128,128,128,0.2)',
        showgrid=True,
        zeroline=False
    )
    
    fig.update_yaxes(
        gridcolor='rgba(128,128,128,0.2)',
        showgrid=True,
        zeroline=False
    )
    
    # Get all trace types in the figure
    trace_types = [trace.type for trace in fig.data]
    
    # If single-color plot, use primary color
    if use_primary_color and len(fig.data) == 1:
        # Handle different trace types
        for trace_type in set(trace_types):
            if trace_type in ['bar', 'histogram']:
                fig.update_traces(marker_color=primary_color, selector=dict(type=trace_type))
            elif trace_type == 'scatter':
                fig.update_traces(
                    marker_color=primary_color,
                    line_color=primary_color,
                    selector=dict(type=trace_type)
                )
            elif trace_type == 'box':
                fig.update_traces(
                    marker_color=primary_color,
                    line_color=primary_color,
                    selector=dict(type=trace_type)
                )
            elif trace_type == 'violin':
                fig.update_traces(
                    marker_color=primary_color,
                    line_color=primary_color,
                    selector=dict(type=trace_type)
                )
    
    return fig

# Initialize session state
if 'data_loaded' not in st.session_state:
    st.session_state.data_loaded = False
if 'model_results' not in st.session_state:
    st.session_state.model_results = {}

st.title("Price Elasticity Analytics Dashboard")

# Sidebar
with st.sidebar:
    col1, col2 = st.columns([8, 1])
    with col1:
        st.header("Data Configuration")
    with col2:
        st.markdown(
            """
            <div title="Click ⋮ in top right → Settings → Theme to customize Streamlit appearance. Use 🎨 Customize Colors below to match chart colors.">
                ℹ️
            </div>
            """, 
            unsafe_allow_html=True
        )
    
    # Or use a help button
    if st.button("ℹ️", help="Click ⋮ in top right → Settings → Theme to customize Streamlit appearance. Use 🎨 Customize Colors below to match chart colors.", key="theme_help"):
        st.info("""
        **Quick Theme Guide:**
        1. Click **⋮** → **Settings** → **Theme**
        2. Choose preset or customize colors
        3. Use **🎨 Customize Colors** below to match charts
        """)
    st.header("Data Configuration")
    
    # Add color picker for runtime theme changes
    with st.expander("🎨 Customize Colors", expanded=False):
        st.write("Override theme colors for this session:")
        custom_color = st.color_picker(
            "Primary Color", 
            value=get_primary_color(),
            key="color_picker"
        )
        
        if custom_color != get_primary_color():
            st.session_state.primary_color_override = custom_color
            st.rerun()
        
        if st.button("Reset to Default"):
            if 'primary_color_override' in st.session_state:
                del st.session_state.primary_color_override
                st.rerun()
    
    st.divider()
    
    data_source = st.radio("Data Source", ["Simulate Data", "Upload Data"])
    
    if data_source == "Simulate Data":
        st.subheader("Simulation Parameters")
        n_skus = st.slider("Number of SKUs", 100, 2000, 500)
        n_years = st.slider("Years of Data", 1, 5, 3)
        
        if st.button("Generate Data"):
            with st.spinner("Simulating data..."):
                transactions, skus, traffic, marketing = simulate_ecommerce_data(
                    n_skus=n_skus, 
                    n_days=365*n_years
                )
                
                # Store in session state
                st.session_state.transactions = transactions
                st.session_state.skus = skus
                st.session_state.traffic = traffic
                st.session_state.marketing = marketing
                st.session_state.data_loaded = True
                
                st.success("Data generated successfully!")
    
    # Replace the else block in your sidebar code with:
    else:  # Upload Data
        st.subheader("Upload Your Data")
        
        # Provide instructions
        with st.expander("📋 Required File Format", expanded=True):
            st.markdown("""
            Please upload 4 CSV files with the following columns:
            
            **1. Transactions File:**
            - `date` (YYYY-MM-DD format)
            - `sku_id` (product identifier)
            - `price` (unit price)
            - `quantity` (units sold)
            - `revenue` (total revenue)
            
            **2. SKUs File:**
            - `sku_id` (product identifier)
            - `category` (product category)
            - `sub_category` (product subcategory)
            - `base_price` (baseline price)
            - `base_elasticity` (optional, default: -1.5)
            - `seasonality_strength` (optional, default: 0.1)
            - `trend_strength` (optional, default: 0.0001)
            
            **3. Traffic File:**
            - `date` (YYYY-MM-DD format)
            - `total_traffic` (total website visits)
            - `unique_visitors` (unique visitors)
            - `conversion_rate` (conversion percentage)
            
            **4. Marketing File:**
            - `date` (YYYY-MM-DD format)
            - `display_spend` (display advertising spend)
            - `search_spend` (search advertising spend)
            - `social_spend` (social media spend)
            - `email_campaigns` (optional, number of campaigns)
            """)
        
        # File uploaders
        st.markdown("### Upload Files")
        
        col1, col2 = st.columns(2)
        
        with col1:
            transactions_file = st.file_uploader(
                "Transactions Data",
                type=['csv'],
                key='transactions_upload',
                help="Upload transaction history CSV"
            )
            
            skus_file = st.file_uploader(
                "SKUs Data",
                type=['csv'],
                key='skus_upload',
                help="Upload product metadata CSV"
            )
        
        with col2:
            traffic_file = st.file_uploader(
                "Traffic Data",
                type=['csv'],
                key='traffic_upload',
                help="Upload website traffic CSV"
            )
            
            marketing_file = st.file_uploader(
                "Marketing Data",
                type=['csv'],
                key='marketing_upload',
                help="Upload marketing spend CSV"
            )
        
        # Process uploaded files
        if all([transactions_file, skus_file, traffic_file, marketing_file]):
            if st.button("Process Uploaded Data", type="primary"):
                with st.spinner("Processing uploaded files..."):
                    transactions, skus, traffic, marketing, errors = process_uploaded_files(
                        transactions_file,
                        skus_file,
                        traffic_file,
                        marketing_file
                    )
                    
                    if errors:
                        st.error("❌ Data validation failed:")
                        for error in errors:
                            st.error(f"• {error}")
                    else:
                        # Store in session state
                        st.session_state.transactions = transactions
                        st.session_state.skus = skus
                        st.session_state.traffic = traffic
                        st.session_state.marketing = marketing
                        st.session_state.data_loaded = True
                        
                        # Show data summary
                        st.success("✅ Data uploaded successfully!")
                        
                        with st.expander("📊 Data Summary", expanded=True):
                            col1, col2 = st.columns(2)
                            with col1:
                                st.metric("Total Transactions", f"{len(transactions):,}")
                                st.metric("Unique SKUs", f"{skus['sku_id'].n_unique():,}")
                            with col2:
                                st.metric("Date Range", f"{transactions['date'].min()} to {transactions['date'].max()}")
                                st.metric("Total Revenue", f"${transactions['revenue'].sum():,.2f}")
        
        else:
            st.info("Please upload all 4 required files to proceed.")
        
        # Option to download sample data
        st.divider()
        
        if st.button("📥 Generate Sample CSV Templates"):
            with st.spinner("Generating sample data..."):
                # Generate small sample data
                sample_trans, sample_skus, sample_traffic, sample_marketing = simulate_ecommerce_data(
                    n_skus=50, 
                    n_days=30
                )
                
                # Convert to CSV
                trans_csv = sample_trans.write_csv()
                skus_csv = sample_skus.write_csv()
                traffic_csv = sample_traffic.write_csv()
                marketing_csv = sample_marketing.write_csv()
                
                # Create download buttons
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label="Download Transactions Sample",
                        data=trans_csv,
                        file_name="sample_transactions.csv",
                        mime="text/csv"
                    )
                    st.download_button(
                        label="Download SKUs Sample",
                        data=skus_csv,
                        file_name="sample_skus.csv",
                        mime="text/csv"
                    )
                with col2:
                    st.download_button(
                        label="Download Traffic Sample",
                        data=traffic_csv,
                        file_name="sample_traffic.csv",
                        mime="text/csv"
                    )
                    st.download_button(
                        label="Download Marketing Sample",
                        data=marketing_csv,
                        file_name="sample_marketing.csv",
                        mime="text/csv"
                    )

# Main content
if st.session_state.data_loaded:
    # Data Processing Tab
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["Raw Data Review", "Processed Data", "Model Comparison", "Results Analysis", "Export"])
    
    with tab1:
        st.header("Raw Data Review")
        
        # Data selection
        data_type = st.selectbox("Select data to review", 
                                ["Transactions", "SKUs", "Traffic", "Marketing"])
        
        if data_type == "Transactions":
            st.subheader("Transaction Data")
            
            # Summary statistics
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Transactions", f"{len(st.session_state.transactions):,}")
            with col2:
                st.metric("Unique SKUs", f"{st.session_state.transactions['sku_id'].n_unique():,}")
            with col3:
                st.metric("Avg Price", f"${st.session_state.transactions['price'].mean():.2f}")
            with col4:
                st.metric("Total Revenue", f"${st.session_state.transactions['revenue'].sum():,.0f}")
            
            # Time series plot
            daily_revenue = st.session_state.transactions.group_by('date').agg([
                pl.col('revenue').sum().alias('total_revenue'),
                pl.col('quantity').sum().alias('total_quantity')
            ]).sort('date').to_pandas()
            
            fig = px.line(daily_revenue, x='date', y='total_revenue', 
                        title="Daily Revenue Trend")
            fig = style_plotly_fig(fig)
            st.plotly_chart(fig, use_container_width=True)
            
            # Sample data with filtering
            st.subheader("Sample Transaction Data")
            
            col1, col2 = st.columns([1, 3])
            with col1:
                sample_size = st.number_input("Sample size", min_value=10, max_value=10000, value=100)
                sample_sku = st.selectbox("Filter by SKU (optional)", 
                                        ["All"] + st.session_state.transactions['sku_id'].unique().to_list()[:50])
            
            with col2:
                if sample_sku == "All":
                    sample_data = st.session_state.transactions.sample(n=min(sample_size, len(st.session_state.transactions)))
                else:
                    sample_data = st.session_state.transactions.filter(pl.col('sku_id') == sample_sku).head(sample_size)
                
                st.dataframe(sample_data.to_pandas())
            
            # Distribution plots
            col1, col2 = st.columns(2)
            with col1:
                fig = px.histogram(st.session_state.transactions.to_pandas(), x='price', 
                                nbins=50, title="Price Distribution")
                fig = style_plotly_fig(fig)
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                fig = px.histogram(st.session_state.transactions.to_pandas(), x='quantity', 
                                 nbins=50, title="Quantity Distribution")
                fig = style_plotly_fig(fig)
                st.plotly_chart(fig, use_container_width=True)
        
        elif data_type == "SKUs":
            st.subheader("SKU Master Data")
            
            # Category breakdown
            category_counts = st.session_state.skus.group_by('category').count().sort('count', descending=True)
            
            col1, col2 = st.columns(2)
            with col1:
                fig = px.bar(category_counts.to_pandas(), x='category', y='count',
                           title="SKUs by Category")
                fig = style_plotly_fig(fig)
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                # Base price by category
                price_by_cat = st.session_state.skus.group_by('category').agg([
                    pl.col('base_price').mean().alias('avg_base_price'),
                    pl.col('base_price').std().alias('std_base_price')
                ]).to_pandas()
                
                fig = px.bar(price_by_cat, x='category', y='avg_base_price',
                           error_y='std_base_price', title="Average Base Price by Category")
                fig = style_plotly_fig(fig)
                st.plotly_chart(fig, use_container_width=True)
            
            # Elasticity distribution
            fig = px.histogram(st.session_state.skus.to_pandas(), x='base_elasticity',
                             nbins=30, title="Base Elasticity Distribution")
            fig = style_plotly_fig(fig)
            st.plotly_chart(fig, use_container_width=True)
            
            # Sample SKU data
            st.subheader("Sample SKU Data")
            st.dataframe(st.session_state.skus.head(100).to_pandas())
        
        elif data_type == "Traffic":
            st.subheader("Traffic Data")
            
            traffic_df = st.session_state.traffic.to_pandas()
            
            # Time series
            fig = px.line(traffic_df, x='date', y='total_traffic',
                         title="Daily Traffic Trend")
            fig = style_plotly_fig(fig)
            st.plotly_chart(fig, use_container_width=True)
            
            # Summary stats
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Avg Daily Traffic", f"{traffic_df['total_traffic'].mean():,.0f}")
            with col2:
                st.metric("Avg Conversion Rate", f"{traffic_df['conversion_rate'].mean():.2%}")
            with col3:
                st.metric("Peak Traffic Day", f"{traffic_df.loc[traffic_df['total_traffic'].idxmax(), 'date']:%Y-%m-%d}")
            
            st.dataframe(traffic_df.head(100))
        
        else:  # Marketing
            st.subheader("Marketing Spend Data")
            
            marketing_df = st.session_state.marketing.to_pandas()
            
            # Calculate total spend
            marketing_df['total_spend'] = marketing_df['display_spend'] + marketing_df['search_spend'] + marketing_df['social_spend']
            
            # Time series of spend by channel
            spend_cols = ['display_spend', 'search_spend', 'social_spend']
            colors = get_current_colors()
            fig = px.line(marketing_df, x='date', y=spend_cols,
                        title="Marketing Spend by Channel Over Time",
                        color_discrete_sequence=colors['palette'][:len(spend_cols)])
            fig = style_plotly_fig(fig, use_primary_color=False)
            st.plotly_chart(fig, use_container_width=True)
            
            # Summary
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Marketing Spend", f"${marketing_df['total_spend'].sum():,.0f}")
            with col2:
                st.metric("Avg Daily Spend", f"${marketing_df['total_spend'].mean():,.0f}")
            with col3:
                st.metric("Total Email Campaigns", f"{marketing_df['email_campaigns'].sum():,}")
            
            st.dataframe(marketing_df.head(100))
    
    with tab2:
        st.header("Processed Data Overview")
        
        # Clean and merge data
        if 'merged_data' not in st.session_state:
            with st.spinner("Processing data..."):
                merged_data = clean_and_merge_data(
                    st.session_state.transactions,
                    st.session_state.traffic,
                    st.session_state.marketing,
                    st.session_state.skus
                )
                st.session_state.merged_data = merged_data
        else:
            merged_data = st.session_state.merged_data
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Records", f"{len(merged_data):,}")
        with col2:
            st.metric("Date Range", f"{merged_data['date'].min()} to {merged_data['date'].max()}")
        with col3:
            st.metric("Unique SKUs", f"{merged_data['sku_id'].n_unique():,}")
        with col4:
            st.metric("Avg Records per SKU", f"{len(merged_data) / merged_data['sku_id'].n_unique():.0f}")
        
        # Data quality check
        st.subheader("Data Quality Check")
        col1, col2 = st.columns(2)
        
        with col1:
            # Check for nulls
            null_counts = {}
            for col in merged_data.columns:
                null_count = merged_data[col].null_count()
                if null_count > 0:
                    null_counts[col] = null_count
            
            if null_counts:
                null_df = pd.DataFrame(list(null_counts.items()), columns=['Column', 'Null Count'])
                st.write("Columns with null values:")
                st.dataframe(null_df)
            else:
                st.success("No null values found!")
        
        with col2:
            # Check log transformations
            st.write("Log transformation check:")
            log_issues = merged_data.filter(
                (pl.col('log_price').is_null()) | 
                (pl.col('log_quantity').is_null()) |
                (pl.col('log_price').is_infinite()) |
                (pl.col('log_quantity').is_infinite())
            )
            st.metric("Records with log issues", len(log_issues))
            if len(log_issues) > 0:
                st.write("Sample of problematic records:")
                st.dataframe(log_issues.head(10).to_pandas())
        
        # Price-Quantity relationship preview
        st.subheader("Price-Quantity Relationship Preview")
        
        # Select a few SKUs to visualize
        sample_skus = merged_data['sku_id'].unique()[:5]
        sample_data = merged_data.filter(pl.col('sku_id').is_in(sample_skus)).to_pandas()
        
        fig = px.scatter(sample_data, x='log_price', y='log_quantity', 
                        color='sku_id', trendline="ols",
                        title="Log Price vs Log Quantity (Sample SKUs)")
        fig = style_plotly_fig(fig, use_primary_color=False)
        st.plotly_chart(fig, use_container_width=True)
        
        # Sample merged data
        st.subheader("Sample Merged Data")
        st.dataframe(merged_data.head(100).to_pandas())
    
    with tab3:
        st.header("Model Comparison")
        
        # Model selection
        col1, col2 = st.columns([1, 3])
        with col1:
            models_to_run = st.multiselect(
                "Select Models",
                ["OLS Log-Log", "Fixed Effects Panel", "Hierarchical Bayes"],
                default=["OLS Log-Log"]
            )
            
            # Sample size for Hierarchical Bayes
            if "Hierarchical Bayes" in models_to_run:
                hb_sample_size = st.number_input(
                    "HB Sample Size (for speed)",
                    min_value=1000,
                    max_value=50000,
                    value=10000,
                    step=1000
                )
            
            if st.button("Run Models"):
                model_runner = PriceElasticityModels()
                
                if "OLS Log-Log" in models_to_run:
                    with st.spinner("Running OLS Log-Log model..."):
                        ols_results = model_runner.ols_log_log(merged_data)
                        st.session_state.model_results['ols'] = ols_results
                
                if "Fixed Effects Panel" in models_to_run:
                    with st.spinner("Running Fixed Effects model..."):
                        fe_results = model_runner.fixed_effects_panel(merged_data)
                        st.session_state.model_results['fixed_effects'] = fe_results
                
                if "Hierarchical Bayes" in models_to_run:
                    with st.spinner("Running Hierarchical Bayes model... (this may take a few minutes)"):
                        sample_size = hb_sample_size if 'hb_sample_size' in locals() else 10000
                        hb_results, trace = model_runner.hierarchical_bayes(
                            merged_data.sample(n=min(sample_size, len(merged_data)))
                        )
                        st.session_state.model_results['hierarchical_bayes'] = hb_results
                
                st.success("Models completed!")
        
        with col2:
            if st.session_state.model_results:
                # Compile results for comparison
                comparison_data = []
                
                for model_name, results in st.session_state.model_results.items():
                    for sku, metrics in results.items():
                        comparison_data.append({
                            'Model': model_name,
                            'SKU': sku,
                            'Elasticity': metrics['elasticity'],
                            'Std_Error': metrics['std_error'],
                            'CI_Lower': metrics['conf_int_lower'],
                            'CI_Upper': metrics['conf_int_upper']
                        })
                
                comparison_df = pd.DataFrame(comparison_data)
                
                # Model comparison visualization
                if len(st.session_state.model_results) > 1:
                    st.subheader("Model Elasticity Comparison")
                    
                    # Select a few SKUs for visualization
                    sample_skus = comparison_df['SKU'].unique()[:10]
                    sample_df = comparison_df[comparison_df['SKU'].isin(sample_skus)]
                    
                    fig = px.scatter(sample_df, x='SKU', y='Elasticity', color='Model',
                                   error_y='Std_Error', title="Elasticity Estimates by Model")
                    fig.update_layout(xaxis_tickangle=-45)
                    fig = style_plotly_fig(fig, use_primary_color=False)
                    st.plotly_chart(fig, use_container_width=True)
                    
                    # Model agreement
                    st.subheader("Model Agreement Analysis")
                    if len(st.session_state.model_results) >= 2:
                        models = list(st.session_state.model_results.keys())[:2]
                        
                        # Get common SKUs
                        skus1 = set(st.session_state.model_results[models[0]].keys())
                        skus2 = set(st.session_state.model_results[models[1]].keys())
                        common_skus = skus1.intersection(skus2)
                        
                        if common_skus:
                            elasticities1 = [st.session_state.model_results[models[0]][sku]['elasticity'] 
                                        for sku in common_skus]
                            elasticities2 = [st.session_state.model_results[models[1]][sku]['elasticity'] 
                                        for sku in common_skus]
                            
                            # Check if either model has constant elasticity (like Fixed Effects)
                            is_constant1 = len(set(elasticities1)) == 1
                            is_constant2 = len(set(elasticities2)) == 1
                            
                            if is_constant1 or is_constant2:
                                # Special handling for models with constant elasticity
                                st.info(f"Note: {'Fixed Effects' if 'fixed_effects' in models else 'One of the models'} estimates a single elasticity for all SKUs.")
                                
                                # Show comparison as a distribution plot instead
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    # Distribution of elasticities
                                    comparison_df = pd.DataFrame({
                                        models[0]: elasticities1,
                                        models[1]: elasticities2,
                                        'SKU': list(common_skus)
                                    })
                                    
                                    # Melt for plotting
                                    melted_df = comparison_df.melt(id_vars=['SKU'], 
                                                                value_vars=models,
                                                                var_name='Model', 
                                                                value_name='Elasticity')
                                    
                                    fig = px.histogram(melted_df, x='Elasticity', color='Model', 
                                                    nbins=30, opacity=0.7,
                                                    title="Distribution of Elasticity Estimates by Model")
                                    fig = style_plotly_fig(fig, use_primary_color=False)

                                    # Add vertical lines for means
                                    for model in models:
                                        model_data = melted_df[melted_df['Model'] == model]
                                        mean_val = model_data['Elasticity'].mean()
                                        fig.add_vline(x=mean_val, 
                                                    line_dash="dash",
                                                    annotation_text=f"{model}: {mean_val:.3f}")
                                    
                                    st.plotly_chart(fig, use_container_width=True)
                                
                                with col2:
                                    # Summary statistics
                                    st.write("**Model Comparison Summary**")
                                    
                                    summary_stats = []
                                    for i, model in enumerate(models):
                                        elasticities = [elasticities1, elasticities2][i]
                                        summary_stats.append({
                                            'Model': model,
                                            'Mean': np.mean(elasticities),
                                            'Std Dev': np.std(elasticities),
                                            'Min': np.min(elasticities),
                                            'Max': np.max(elasticities),
                                            'Unique Values': len(set(elasticities))
                                        })
                                    
                                    summary_df = pd.DataFrame(summary_stats)
                                    st.dataframe(summary_df.style.format({
                                        'Mean': '{:.3f}',
                                        'Std Dev': '{:.3f}',
                                        'Min': '{:.3f}',
                                        'Max': '{:.3f}'
                                    }))
                                    
                                    # If one is constant, show how much OLS varies around the FE estimate
                                    if is_constant1 and not is_constant2:
                                        fe_estimate = elasticities1[0]
                                        ols_deviation = [abs(e - fe_estimate) for e in elasticities2]
                                        st.metric("Mean Absolute Deviation from FE estimate", 
                                                f"{np.mean(ols_deviation):.3f}")
                                    elif is_constant2 and not is_constant1:
                                        fe_estimate = elasticities2[0]
                                        ols_deviation = [abs(e - fe_estimate) for e in elasticities1]
                                        st.metric("Mean Absolute Deviation from FE estimate", 
                                                f"{np.mean(ols_deviation):.3f}")
                            
                            else:
                                # Standard scatter plot for models with varying elasticities
                                colors = get_current_colors()
                                fig = px.scatter(x=elasticities1, y=elasticities2,
                                            labels={'x': f'{models[0]} Elasticity',
                                                    'y': f'{models[1]} Elasticity'},
                                            title=f"Model Agreement: {models[0]} vs {models[1]}",
                                            color_discrete_sequence=colors['palette'])
                                fig = style_plotly_fig(fig, use_primary_color=True)
                                
                                # Add 45-degree line
                                min_val = min(min(elasticities1), min(elasticities2))
                                max_val = max(max(elasticities1), max(elasticities2))
                                fig.add_trace(go.Scatter(x=[min_val, max_val], y=[min_val, max_val],
                                                    mode='lines', name='Perfect Agreement',
                                                    line=dict(dash='dash')))
                                
                                st.plotly_chart(fig, use_container_width=True)
                                
                                # Calculate correlation
                                correlation = np.corrcoef(elasticities1, elasticities2)[0, 1]
                                st.metric("Correlation between models", f"{correlation:.3f}")
                            
                            # Always show detailed comparison table
                            st.subheader("Detailed Model Comparison")
                            
                            # Create comparison dataframe
                            detailed_comparison = []
                            for sku in list(common_skus)[:20]:  # Show first 20 SKUs
                                row = {'SKU': sku}
                                for model in models:
                                    if sku in st.session_state.model_results[model]:
                                        metrics = st.session_state.model_results[model][sku]
                                        row[f'{model}_elasticity'] = metrics['elasticity']
                                        row[f'{model}_CI_width'] = metrics['conf_int_upper'] - metrics['conf_int_lower']
                                detailed_comparison.append(row)
                            
                            detail_df = pd.DataFrame(detailed_comparison)
                            
                            # Add difference column if both models have varying elasticities
                            if not is_constant1 and not is_constant2:
                                detail_df['Difference'] = detail_df[f'{models[0]}_elasticity'] - detail_df[f'{models[1]}_elasticity']
                                detail_df['Abs_Difference'] = abs(detail_df['Difference'])
                                detail_df = detail_df.sort_values('Abs_Difference', ascending=False)
                            
                            st.dataframe(detail_df.style.format({
                                col: '{:.3f}' for col in detail_df.columns if 'elasticity' in col or 'Difference' in col or 'CI_width' in col
                            }))
    
    with tab4:
        st.header("Results Analysis")
        
        if st.session_state.model_results:
            # Let user select which model results to analyze
            available_models = list(st.session_state.model_results.keys())
            
            if len(available_models) == 1:
                selected_model = available_models[0]
            else:
                selected_model = st.selectbox("Select Model to Analyze", available_models)
            
            if selected_model and selected_model in st.session_state.model_results:
                results = st.session_state.model_results[selected_model]
                
                # Convert to DataFrame for analysis
                results_list = []
                for sku, metrics in results.items():
                    sku_info = st.session_state.skus.filter(pl.col('sku_id') == sku)
                    if len(sku_info) > 0:
                        results_list.append({
                            'SKU': sku,
                            'Category': sku_info['category'][0],
                            'Sub_Category': sku_info['sub_category'][0],
                            'Elasticity': metrics['elasticity'],
                            'CI_Lower': metrics['conf_int_lower'],
                            'CI_Upper': metrics['conf_int_upper'],
                            'P_Value': metrics.get('p_value', 0),
                            'R_Squared': metrics.get('r_squared', 0)
                        })
                
                results_df = pd.DataFrame(results_list)
                
                if len(results_df) > 0:
                    # Elasticity buckets
                    def categorize_elasticity(elasticity):
                        if elasticity > -0.5:
                            return "Inelastic"
                        elif elasticity > -1.0:
                            return "Unit Elastic"
                        elif elasticity > -1.5:
                            return "Elastic"
                        else:
                            return "Very Elastic"
                    
                    results_df['Elasticity_Bucket'] = results_df['Elasticity'].apply(categorize_elasticity)
                    
                    # Display model name
                    st.subheader(f"Analysis for {selected_model.replace('_', ' ').title()} Model")
                    
                    # Display summary statistics
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.subheader("Elasticity Distribution")
                        fig = px.histogram(results_df, x='Elasticity', nbins=30,
                                        title="Distribution of Price Elasticities")
                        fig.add_vline(x=results_df['Elasticity'].mean(), line_dash="dash", 
                                    annotation_text=f"Mean: {results_df['Elasticity'].mean():.2f}")
                        fig = style_plotly_fig(fig, use_primary_color=True)
                        st.plotly_chart(fig, use_container_width=True)
                        
                        st.subheader("Elasticity by Category")
                        category_summary = results_df.groupby('Category')['Elasticity'].agg(['mean', 'std', 'count']).reset_index()
                        category_summary.columns = ['Category', 'Mean Elasticity', 'Std Dev', 'Count']
                        st.dataframe(category_summary.style.format({
                            'Mean Elasticity': '{:.3f}',
                            'Std Dev': '{:.3f}'
                        }))
                    
                    with col2:
                        st.subheader("Elasticity Buckets")
                        bucket_counts = results_df['Elasticity_Bucket'].value_counts()
                        colors = get_current_colors()
                        fig = px.pie(values=bucket_counts.values, names=bucket_counts.index,
                                    title="SKUs by Elasticity Bucket",
                                    color_discrete_sequence=colors['palette'])
                        fig = style_plotly_fig(fig, use_primary_color=False)
                        st.plotly_chart(fig, use_container_width=True)
                        
                        st.subheader("Statistical Significance")
                        if 'P_Value' in results_df.columns and results_df['P_Value'].sum() > 0:
                            sig_df = results_df[results_df['P_Value'] < 0.05]
                            st.metric("Statistically Significant Results", f"{len(sig_df)} / {len(results_df)}")
                            
                            # Show confidence interval width
                            results_df['CI_Width'] = results_df['CI_Upper'] - results_df['CI_Lower']
                            avg_ci_width = results_df['CI_Width'].mean()
                            st.metric("Average Confidence Interval Width", f"{avg_ci_width:.3f}")
                        else:
                            st.info("P-values not available for this model")
                    
                    # Additional visualizations
                    st.subheader("Elasticity Analysis by Category")
                    
                    # Box plot by category
                    fig = px.box(results_df, x='Category', y='Elasticity',
                                title="Elasticity Distribution by Category")
                    fig = style_plotly_fig(fig, use_primary_color=True)
                    fig.update_layout(xaxis_tickangle=-45)
                    st.plotly_chart(fig, use_container_width=True)
                                        
                    # Detailed results table with filtering
                    st.subheader("Detailed Results")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        category_filter = st.multiselect("Filter by Category", 
                                                    results_df['Category'].unique(),
                                                    default=results_df['Category'].unique())
                    with col2:
                        bucket_filter = st.multiselect("Filter by Elasticity Bucket",
                                                    results_df['Elasticity_Bucket'].unique(),
                                                    default=results_df['Elasticity_Bucket'].unique())
                    with col3:
                        if results_df['R_Squared'].sum() > 0:  # Check if R-squared values exist
                            min_r2 = st.slider("Minimum R-squared", 0.0, 1.0, 0.0)
                        else:
                            min_r2 = 0.0
                            st.info("R-squared not available for this model")
                    
                    filtered_df = results_df[
                        (results_df['Category'].isin(category_filter)) &
                        (results_df['Elasticity_Bucket'].isin(bucket_filter)) &
                        (results_df['R_Squared'] >= min_r2)
                    ]
                    
                    # Sort by elasticity
                    filtered_df = filtered_df.sort_values('Elasticity')
                    
                    # Display with formatting
                    display_df = filtered_df[['SKU', 'Category', 'Sub_Category', 'Elasticity', 
                                            'CI_Lower', 'CI_Upper', 'Elasticity_Bucket']]
                    
                    if results_df['R_Squared'].sum() > 0:
                        display_df = filtered_df[['SKU', 'Category', 'Sub_Category', 'Elasticity', 
                                                'CI_Lower', 'CI_Upper', 'R_Squared', 'Elasticity_Bucket']]
                    
                    st.dataframe(
                        display_df.style.format({
                            'Elasticity': '{:.3f}',
                            'CI_Lower': '{:.3f}',
                            'CI_Upper': '{:.3f}',
                            'R_Squared': '{:.3f}'
                        }).background_gradient(subset=['Elasticity'], cmap='RdYlBu_r')
                    )
                    
                    # Summary metrics
                    st.subheader("Summary Metrics")
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Mean Elasticity", f"{filtered_df['Elasticity'].mean():.3f}")
                    with col2:
                        st.metric("Median Elasticity", f"{filtered_df['Elasticity'].median():.3f}")
                    with col3:
                        st.metric("Most Common Bucket", filtered_df['Elasticity_Bucket'].mode()[0])
                    with col4:
                        st.metric("SKUs Analyzed", len(filtered_df))
                    
                else:
                    st.warning("No results available for analysis")
        else:
            st.info("Please run models to generate results for analysis")
    
    with tab5:
        st.header("Export Results")
        
        if st.session_state.model_results:
            # Prepare export data
            export_data = []
            
            def categorize_elasticity(elasticity):
                if elasticity > -0.5:
                    return "Inelastic"
                elif elasticity > -1.0:
                    return "Unit Elastic"
                elif elasticity > -1.5:
                    return "Elastic"
                else:
                    return "Very Elastic"
            
            for model_name, results in st.session_state.model_results.items():
                if not results:  # Skip empty results
                    continue
                    
                for sku, metrics in results.items():
                    sku_info = st.session_state.skus.filter(pl.col('sku_id') == sku)
                    if len(sku_info) > 0:
                        row = {
                            'Model': model_name,
                            'SKU': sku,
                            'Category': sku_info['category'][0],
                            'Sub_Category': sku_info['sub_category'][0],
                            'Elasticity': metrics['elasticity'],
                            'Std_Error': metrics['std_error'],
                            'CI_Lower': metrics['conf_int_lower'],
                            'CI_Upper': metrics['conf_int_upper'],
                            'Elasticity_Bucket': categorize_elasticity(metrics['elasticity'])
                        }
                        
                        # Add model-specific metrics
                        if 'r_squared' in metrics:
                            row['R_Squared'] = metrics['r_squared']
                        if 'p_value' in metrics:
                            row['P_Value'] = metrics['p_value']
                        if 'entity_effect' in metrics:
                            row['Entity_Effect'] = metrics['entity_effect']
                        
                        export_data.append(row)
            
            if export_data:  # Only proceed if we have data
                export_df = pd.DataFrame(export_data)
                
                st.write(f"Total results to export: {len(export_df)}")
                
                # Debug: Show columns
                st.write("Available columns:", export_df.columns.tolist())
                
                # Category rollups - with error handling
                try:
                    category_rollup = export_df.groupby(['Model', 'Category'])['Elasticity'].agg(['mean', 'std', 'count']).reset_index()
                    category_rollup.columns = ['Model', 'Category', 'Mean_Elasticity', 'Std_Elasticity', 'Count']
                except KeyError as e:
                    st.error(f"Error creating category rollup: {e}")
                    st.write("Export DataFrame shape:", export_df.shape)
                    st.write("Export DataFrame head:", export_df.head())
                    category_rollup = pd.DataFrame()
                
                try:
                    sub_category_rollup = export_df.groupby(['Model', 'Sub_Category'])['Elasticity'].agg(['mean', 'std', 'count']).reset_index()
                    sub_category_rollup.columns = ['Model', 'Sub_Category', 'Mean_Elasticity', 'Std_Elasticity', 'Count']
                except KeyError as e:
                    st.error(f"Error creating sub-category rollup: {e}")
                    sub_category_rollup = pd.DataFrame()
                
                # Create Excel file with formatting
                output = io.BytesIO()
                try:
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # Write main results
                        export_df.to_excel(writer, sheet_name='SKU_Results', index=False)
                        if not category_rollup.empty:
                            category_rollup.to_excel(writer, sheet_name='Category_Rollup', index=False)
                        if not sub_category_rollup.empty:
                            sub_category_rollup.to_excel(writer, sheet_name='SubCategory_Rollup', index=False)
                        
                        # Get the workbook and worksheets
                        workbook = writer.book
                        worksheet = writer.sheets['SKU_Results']
                        
                        # Define formatting styles
                        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                        from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
                        
                        # Header formatting
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Find column indices
                        headers = [cell.value for cell in worksheet[1]]
                        elasticity_col = headers.index('Elasticity') + 1 if 'Elasticity' in headers else None
                        bucket_col = headers.index('Elasticity_Bucket') + 1 if 'Elasticity_Bucket' in headers else None
                        ci_lower_col = headers.index('CI_Lower') + 1 if 'CI_Lower' in headers else None
                        ci_upper_col = headers.index('CI_Upper') + 1 if 'CI_Upper' in headers else None
                        r_squared_col = headers.index('R_Squared') + 1 if 'R_Squared' in headers else None
                        
                        # Apply conditional formatting
                        last_row = worksheet.max_row
                        
                        if elasticity_col:
                            # Color scale for elasticity values
                            # Red (inelastic) -> Yellow (unit elastic) -> Green (very elastic)
                            elasticity_range = f'{chr(64 + elasticity_col)}2:{chr(64 + elasticity_col)}{last_row}'
                            
                            # Create custom color scale
                            color_scale_rule = ColorScaleRule(
                                start_type='num', start_value=-3, start_color='00FF00',  # Green for very elastic
                                mid_type='num', mid_value=-1, mid_color='FFFF00',        # Yellow for unit elastic
                                end_type='num', end_value=0, end_color='FF0000'          # Red for inelastic
                            )
                            worksheet.conditional_formatting.add(elasticity_range, color_scale_rule)
                        
                        if bucket_col:
                            # Highlight elasticity buckets with different colors
                            bucket_range = f'{chr(64 +bucket_col)}2:{chr(64 + bucket_col)}{last_row}'
                            
                            # Define fills for each bucket
                            very_elastic_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                            elastic_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
                            unit_elastic_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                            inelastic_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
                            
                            # Apply formatting based on bucket value
                            for row in range(2, last_row + 1):
                                cell = worksheet.cell(row=row, column=bucket_col)
                                if cell.value == "Very Elastic":
                                    cell.fill = very_elastic_fill
                                elif cell.value == "Elastic":
                                    cell.fill = elastic_fill
                                elif cell.value == "Unit Elastic":
                                    cell.fill = unit_elastic_fill
                                elif cell.value == "Inelastic":
                                    cell.fill = inelastic_fill
                        
                        if ci_lower_col and ci_upper_col:
                            # Calculate and highlight confidence interval width
                            for row in range(2, last_row + 1):
                                ci_lower = worksheet.cell(row=row, column=ci_lower_col).value
                                ci_upper = worksheet.cell(row=row, column=ci_upper_col).value
                                
                                if ci_lower is not None and ci_upper is not None:
                                    ci_width = ci_upper - ci_lower
                                    
                                    # Highlight wide confidence intervals (low precision)
                                    if ci_width > 1.0:  # If CI width > 1
                                        for col in [ci_lower_col, ci_upper_col]:
                                            cell = worksheet.cell(row=row, column=col)
                                            cell.fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
                                            cell.font = Font(color="C00000")
                        
                        if r_squared_col:
                            # Highlight R-squared values
                            r_squared_range = f'{chr(64 + r_squared_col)}2:{chr(64 + r_squared_col)}{last_row}'
                            
                            # Poor fit (R² < 0.3) in red
                            poor_fit_rule = CellIsRule(operator='lessThan', formula=['0.3'], 
                                                    fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                                                    font=Font(color="9C0006"))
                            
                            # Good fit (R² > 0.7) in green
                            good_fit_rule = CellIsRule(operator='greaterThan', formula=['0.7'],
                                                    fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                                                    font=Font(color="006100"))
                            
                            worksheet.conditional_formatting.add(r_squared_range, poor_fit_rule)
                            worksheet.conditional_formatting.add(r_squared_range, good_fit_rule)
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # Freeze the header row
                        worksheet.freeze_panes = 'A2'
                        
                        # Add borders
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        for row in worksheet.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=worksheet.max_column):
                            for cell in row:
                                cell.border = thin_border
                        
                        # Format numeric columns
                        for row in range(2, last_row + 1):
                            if elasticity_col:
                                worksheet.cell(row=row, column=elasticity_col).number_format = '0.000'
                            if ci_lower_col:
                                worksheet.cell(row=row, column=ci_lower_col).number_format = '0.000'
                            if ci_upper_col:
                                worksheet.cell(row=row, column=ci_upper_col).number_format = '0.000'
                            if r_squared_col:
                                worksheet.cell(row=row, column=r_squared_col).number_format = '0.000'
                        
                        # Add a summary sheet with insights
                        summary_sheet = workbook.create_sheet('Summary')
                        
                        # Summary statistics
                        summary_data = [
                            ['Elasticity Analysis Summary', ''],
                            ['', ''],
                            ['Metric', 'Value'],
                            ['Total SKUs Analyzed', len(export_df)],
                            ['Models Run', ', '.join(export_df['Model'].unique())],
                            ['', ''],
                            ['Elasticity Distribution', ''],
                            ['Mean Elasticity', f"{export_df['Elasticity'].mean():.3f}"],
                            ['Median Elasticity', f"{export_df['Elasticity'].median():.3f}"],
                            ['Std Dev', f"{export_df['Elasticity'].std():.3f}"],
                            ['', ''],
                            ['Elasticity Buckets', ''],
                        ]
                        
                        # Add bucket counts
                        bucket_counts = export_df['Elasticity_Bucket'].value_counts()
                        for bucket, count in bucket_counts.items():
                            summary_data.append([bucket, f"{count} ({count/len(export_df)*100:.1f}%)"])
                        
                        # Write summary data
                        for row_idx, row_data in enumerate(summary_data, 1):
                            for col_idx, value in enumerate(row_data, 1):
                                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                                
                                # Format headers
                                if row_idx in [1, 3, 7, 12]:
                                    cell.font = Font(bold=True, size=12)
                                    if row_idx == 1:
                                        cell.font = Font(bold=True, size=14)
                        
                        # Auto-adjust summary sheet columns
                        for column in summary_sheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            summary_sheet.column_dimensions[column_letter].width = adjusted_width
                    
                    output.seek(0)
                    
                    st.download_button(
                        label="Download Excel Report",
                        data=output,
                        file_name=f"price_elasticity_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.success("Report ready for download!")
                    
                    # Preview
                    st.subheader("Export Preview")
                    st.dataframe(export_df.head(20))
                    
                    # Show summary statistics
                    col1, col2 = st.columns(2)
                    with col1:
                        st.subheader("Results Summary")
                        summary_stats = export_df.groupby('Model')['Elasticity'].describe()
                        st.dataframe(summary_stats)
                    
                    with col2:
                        st.subheader("Elasticity Buckets by Model")
                        bucket_summary = export_df.groupby(['Model', 'Elasticity_Bucket']).size().unstack(fill_value=0)
                        st.dataframe(bucket_summary)
                        
                except Exception as e:
                    st.error(f"Error creating Excel file: {e}")
                    st.write("Export data sample:", export_df.head() if not export_df.empty else "No data")
                    
            else:
                st.warning("No results available to export. Please run models first.")
        else:
            st.info("Please run models to generate results for export")
else:
    st.info("Please generate or upload data to begin analysis")